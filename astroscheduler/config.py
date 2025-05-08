from datetime import datetime
import os
import shutil
import openpyxl
import json

class AstroSchedulerConfig:
    def __init__(self, config_dir=None):
        """
        Initialize the AstroSchedulerConfig class.

        :param config_dir: Directory where configuration files are stored.
        """
        # Default configuration directory
        self.config_dir = config_dir or os.path.join(os.path.dirname(__file__), "config")
        self.template_file = os.path.join(self.config_dir, "TimeScheduleConfig.xlsx")

        # Ensure the configuration directory exists
        os.makedirs(self.config_dir, exist_ok=True)

        # Map of expected keys to instance attributes
        self.config_keys = {
            "Latitude": "latitude",
            "Longitude": "longitude",
            "ScheduleType": "schedule_type",
            "DefaultValue": "default_value",
            "ReferenceYear": "reference_year",
            "EBOVersion": "ebo_version",
            "ScheduleName": "schedule_name",
        }
        # Initialize default values for configuration attributes
        self.latitude = None
        self.longitude = None
        self.schedule_type = "Multistate"
        self.default_value = 0
        self.reference_year = 2025
        self.ebo_version = "4.0.1"
        self.schedule_name = "AstroScheduler"

        # Initialize entries as an empty list
        self.entries = []


    def copy_sample_template(self, destination_path, new_name=None):
        """
        Copy the sample Excel configuration template to the provided location.
        Optionally rename the copied template file.

        :param destination_path: Path where the template should be copied.
        :param new_name: Optional new name for the copied template file.
        :raises FileNotFoundError: If the template file does not exist.
        :raises ValueError: If the destination path is invalid.
        """
        if not os.path.exists(self.template_file):
            raise FileNotFoundError(f"Template file not found: {self.template_file}")

        # Ensure the destination directory exists
        destination_dir = os.path.dirname(destination_path)
        if not destination_dir:
            raise ValueError("Invalid destination path provided.")

        os.makedirs(destination_dir, exist_ok=True)

        # Determine the final destination file path
        if new_name:
            destination_path = os.path.join(destination_dir, new_name)

        # Copy the template file to the destination
        shutil.copy(self.template_file, destination_path)
        print(f"Template copied to: {destination_path}")

    def from_spreadsheet(self, file_path):
        """
        Load the spreadsheet into a dictionary. The 'Entries' sheet will be loaded
        into 'self.entries' and contain an array of dicts where the first row is the keys
        and subsequent rows are the values. The 'Configuration' sheet will be loaded
        to set specific attributes like latitude, longitude, schedule_type, etc.

        :param file_path: Path to the Excel file to load.
        :raises FileNotFoundError: If the file does not exist.
        :raises ValueError: If required sheets or keys are not found.
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        # Load the workbook
        workbook = openpyxl.load_workbook(file_path)

        # Load the 'Entries' sheet
        if 'Entries' not in workbook.sheetnames:
            raise ValueError("The 'Entries' sheet is not found in the spreadsheet.")

        sheet_entries = workbook['Entries']
        entries_data = []

        # Extract keys from the first row
        keys = [cell.value for cell in sheet_entries[1]]

        # Check for required keys
        required_keys = ["TimeReference", "Hour", "Minute", "Value"]
        missing_keys = [key for key in required_keys if key not in keys]
        if missing_keys:
            raise ValueError(f"The following required keys are missing in the 'Entries' sheet: {', '.join(missing_keys)}")

        # Extract values from subsequent rows
        for row in sheet_entries.iter_rows(min_row=2, values_only=True):
            entry = {keys[i]: row[i] for i in range(len(keys))}
            # Check the Value field for equivalent of None/Null value using the helper function
            entry["Value"] = self.check_for_null_value(entry.get("Value"))
            entries_data.append(entry)

        # Save the data to self.entries
        self.entries = entries_data

        # Load the 'Configuration' sheet
        if 'Configuration' not in workbook.sheetnames:
            raise ValueError("The 'Configuration' sheet is not found in the spreadsheet.")

        sheet_config = workbook['Configuration']

        # Iterate through rows in the Configuration sheet
        for row in sheet_config.iter_rows(min_row=1, max_col=2, values_only=True):
            key, value = row
            if key is None:
                continue

            # Check if the key is in the expected keys map
            if key in self.config_keys:
                setattr(self, self.config_keys[key], value)

            # Determine the year to use
            year = datetime.now().year
            if self.reference_year is None:
                try:
                    year = int(self.reference_year)
                except (AttributeError, ValueError) as e:
                    year = datetime.now().year  # Default to the current year
                    print(f'Failed to parse config ReferenceYear {self.reference_year}, got error {e}. Using current year {year} instead...')
            self.reference_year = year

    def check_for_null_value(self, value):
        """
        Helper function to check if a value is considered "null" and return None.
        If the value is not null, it attempts to convert it to an integer.

        :param value: The value to check.
        :return: None if the value is considered null, otherwise the integer value.
        :raises ValueError: If the value cannot be converted to an integer.
        """
        if value is None or (isinstance(value, str) and value.strip().lower() in ["null", "none", ""]):
            return None
        try:
            return int(value)
        except ValueError:
            raise ValueError(f"Invalid value: {value}. Must be an integer or one of ['null', 'none', '', None].")

    def to_dict(self):
        """
        Convert the configuration object into a dictionary.

        :return: A dictionary representation of the configuration object.
        """
        return {
            "latitude": self.latitude,
            "longitude": self.longitude,
            "schedule_type": self.schedule_type,
            "default_value": self.default_value,
            "reference_year": self.reference_year,
            "ebo_version": self.ebo_version,
            "entries": self.entries,
        }
    
    def to_json(self, output_file):
        """
        Write the configuration object as a JSON file.

        :param output_file: Path to the output JSON file.
        :raises ValueError: If the output file path is invalid.
        """
        if not output_file or not isinstance(output_file, str):
            raise ValueError("Invalid output file path provided.")

        # Convert the configuration object to a dictionary
        config_dict = self.to_dict()

        # Write the dictionary to a JSON file
        with open(output_file, "w") as json_file:
            json.dump(config_dict, json_file, indent=4)

        print(f"Configuration written to JSON file: {output_file}")

    def add_entry(self, time_ref="Absolute", hour=0, minute=0, value=None):
        """
        Add an entry to the configuration's entries list.

        :param time_ref: The time reference for the entry. Can be "SunriseOffset", "SunsetOffset", or "Absolute" (default).
        :param hour: The hour value for the entry. Can be a string or number, but will be converted to a number.
        :param minute: The minute value for the entry. Can be a string or number, but will be converted to a number.
        :param value: The value for the entry. Can be a string, number, or None.
        :raises ValueError: If time_ref is not one of the allowed values.
        """
        # Validate time_ref
        allowed_time_refs = ["SunriseOffset", "SunsetOffset", "Absolute"]
        if time_ref not in allowed_time_refs:
            raise ValueError(f"time_ref must be one of {allowed_time_refs}, got '{time_ref}'.")

        # Convert hour and minute to numbers
        try:
            hour = int(hour)
        except ValueError:
            raise ValueError(f"hour must be a number or a string that can be converted to a number, got '{hour}'.")

        try:
            minute = int(minute)
        except ValueError:
            raise ValueError(f"minute must be a number or a string that can be converted to a number, got '{minute}'.")

        # Convert value to number or None
        if value is not None:
            try:
                value = int(value)
            except ValueError:
                raise ValueError(f"value must be a number, a string that can be converted to a number, or None, got '{value}'.")

        # Create the entry
        entry = {
            "TimeReference": time_ref,
            "Hour": hour,
            "Minute": minute,
            "Value": value,
        }

        # Add the entry to self.entries
        self.entries.append(entry)        

if __name__ == "__main__":
    # Initialize the configuration class
    config = AstroSchedulerConfig()

    # Define paths for input and output
    script_dir = os.path.dirname(__file__)
    input_dir = os.path.join(script_dir, "../data/sample_input")
    output_dir = os.path.join(script_dir, "../data/sample_output")
    os.makedirs(input_dir, exist_ok=True)
    os.makedirs(output_dir, exist_ok=True)

    # Copy a sample Excel template to the input directory
    template_path = os.path.join(input_dir, "TimeScheduleConfig.xlsx")
    config.copy_sample_template(template_path)

    # Load configuration from the Excel file
    config.from_spreadsheet(template_path)

    # Print configuration attributes
    print("Configuration Attributes:")
    print(f"Latitude: {config.latitude}")
    print(f"Longitude: {config.longitude}")
    print(f"Schedule Type: {config.schedule_type}")
    print(f"Default Value: {config.default_value}")
    print(f"Reference Year: {config.reference_year}")
    print(f"EBO Version: {config.ebo_version}")
    print(f"Entries: {config.entries}")

    # Export configuration to a JSON file
    output_json = os.path.join(output_dir, "config.json")
    config.to_json(output_json)        